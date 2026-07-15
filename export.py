"""
Excel export for NewsIntel.

Builds a formatted .xlsx workbook from the dashboard's view model:
  - "Articles" sheet: one row per article, all fields the dashboard shows
  - "Summary" sheet: run metadata, category breakdown, source health

Pure data export — no formulas, since this is a point-in-time snapshot of a
pipeline run, not a model meant to recalculate.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

FONT_NAME = "Arial"

HEADER_FILL = PatternFill(start_color="0B1220", end_color="0B1220", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="0B1220")
LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True, color="475569")
BODY_FONT = Font(name=FONT_NAME, size=10, color="1E293B")
THIN_BORDER = Border(bottom=Side(style="thin", color="E5E9F0"))


def _style_header_row(ws: Worksheet, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=False)


def _autofit_columns(ws: Worksheet, widths: dict[int, int]):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _build_articles_sheet(wb: Workbook, vm: dict):
    ws = wb.active
    ws.title = "Articles"

    headers = [
        "Title", "URL", "Source", "Outlet Domain", "Published At",
        "Category", "Secondary Categories", "Confidence %",
        "Description", "Categorized By",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for a in vm["articles"]:
        published = a.get("published_at", "")
        ws.append([
            a.get("title", ""),
            a.get("url", ""),
            a.get("source", ""),
            a.get("domain", ""),
            published,
            a.get("category", ""),
            ", ".join(a.get("secondary_categories", []) or []),
            a.get("confidence_pct", 0),
            a.get("description", ""),
            a.get("categorized_by", ""),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column in (1, 9)))

    _autofit_columns(ws, {
        1: 48,   # Title
        2: 42,   # URL
        3: 16,   # Source
        4: 22,   # Domain
        5: 22,   # Published
        6: 26,   # Category
        7: 26,   # Secondary categories
        8: 12,   # Confidence
        9: 60,   # Description
        10: 16,  # Categorized by
    })

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30

    return ws


def _build_summary_sheet(wb: Workbook, vm: dict):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40

    r = 1
    ws.cell(row=r, column=1, value="NewsIntel — Run Summary").font = TITLE_FONT
    r += 2

    meta_rows = [
        ("Query", vm["query"]),
        ("Date range", f'{vm["date_from"]} to {vm["date_to"]}'),
        ("Generated at", vm["generated_at_display"]),
        ("Requested sources", ", ".join(vm["requested_sources"]) if vm["requested_sources"] else "—"),
        ("Total articles", vm["article_count"]),
        ("Average confidence", f'{vm["avg_confidence"]}%'),
        ("Sources healthy", f'{vm["sources_ok_count"]}/{vm["sources_total_count"]}'),
    ]
    for label, value in meta_rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Category Breakdown").font = Font(
        name=FONT_NAME, size=12, bold=True, color="0B1220"
    )
    r += 1
    cat_header_row = r
    ws.cell(row=r, column=1, value="Category")
    ws.cell(row=r, column=2, value="Articles")
    ws.cell(row=r, column=3, value="% of Total")
    _style_header_row(ws, cat_header_row, 3)
    r += 1
    for c in vm["categories"]:
        ws.cell(row=r, column=1, value=c["name"]).font = BODY_FONT
        ws.cell(row=r, column=2, value=c["count"]).font = BODY_FONT
        pct_cell = ws.cell(row=r, column=3, value=c["pct"] / 100)
        pct_cell.font = BODY_FONT
        pct_cell.number_format = "0.0%"
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Source Health").font = Font(
        name=FONT_NAME, size=12, bold=True, color="0B1220"
    )
    r += 1
    src_header_row = r
    ws.cell(row=r, column=1, value="Source")
    ws.cell(row=r, column=2, value="Status")
    ws.cell(row=r, column=3, value="Article Count")
    ws.cell(row=r, column=4, value="Duration (s)")
    _style_header_row(ws, src_header_row, 4)
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    r += 1
    for s in vm["sources"]:
        ws.cell(row=r, column=1, value=s.get("name", "")).font = BODY_FONT
        ws.cell(row=r, column=2, value=s.get("status_label", "")).font = BODY_FONT
        ws.cell(row=r, column=3, value=s.get("article_count", 0)).font = BODY_FONT
        ws.cell(row=r, column=4, value=s.get("duration_seconds", 0)).font = BODY_FONT
        r += 1

    return ws


def build_workbook(vm: dict) -> BytesIO:
    """Build the export workbook in memory and return a seekable BytesIO."""
    wb = Workbook()
    _build_articles_sheet(wb, vm)
    _build_summary_sheet(wb, vm)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
